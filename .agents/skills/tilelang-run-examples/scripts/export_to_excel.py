#!/usr/bin/env python3
"""Export TileLang-Ascend test results to Excel with multi-round comparison.

Usage:
  python export_to_excel.py --log <log_file> --backend <auto|ascendc|pto> [--excel <output_path>]

This script parses the run_examples.sh output log, extracts per-script results,
and exports them to an Excel file. When run multiple times, each round's results
are appended as a new sheet, and a cross-round comparison sheet is automatically
maintained.

Excel file structure (after multiple rounds):
  - "Round 1 (auto/pto)" : Detailed results for round 1
  - "Round 2 (auto/pto)" : Detailed results for round 2
  - ...
  - "失败分类汇总"        : Failure type summary for latest round
  - "对比分析"           : Cross-round comparison (R1 vs R2 vs ...)
"""

import argparse
import os
import re
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=12, color="FFFFFF")
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
fixed_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
new_fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
changed_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def classify_failure(last_lines):
    """Classify the failure type from the last few lines of output."""
    text = "\n".join(last_lines)
    if "Compilation Failed!" in text:
        return "编译失败", "bisheng 编译报错"
    if "Unsupport SyncAll in pto backend" in text:
        return "pto不支持", "Unsupport SyncAll in pto backend"
    if "Unresolved call Op(tl.ascend_reinterpretcast)" in text:
        return "pto不支持", "Unresolved call Op(tl.ascend_reinterpretcast)"
    if "Downcast" in text and "failed" in text:
        return "内部错误", "Downcast type mismatch"
    if "Mismatched elements" in text:
        match = re.search(r"Mismatched elements: (.+?)$", text, re.MULTILINE)
        detail = match.group(1).strip() if match else "精度不匹配"
        return "精度不匹配", detail
    if "accuracy:" in text or "The precision is not correct" in text:
        match = re.search(r"accuracy: ([\d.]+)", text)
        detail = f"accuracy {match.group(1)}" if match else "精度不匹配"
        return "精度不匹配", detail
    if "vector::reserve" in text or "length_error" in text:
        return "NPU设备错误", "std::length_error: vector::reserve"
    if "aicore exception" in text or "rtDeviceSynchronizeWithTimeout" in text:
        return "NPU设备错误", "aicore exception / npuSynchronizeDevice failed"
    if "open device" in text and "failed" in text:
        return "NPU设备错误", "open device failed"
    if "Exit code 139" in text or "Exit: 139" in text:
        return "段错误(Segfault)", "Exit code 139"
    if "Exit code" in text:
        return "运行时错误", text.split("Exit")[-1].strip()[:60]
    return "未知", text[:80]


def parse_log(log_path):
    """Parse run_examples.sh output log and extract per-script results.

    Returns list of (script, status, fail_type, fail_detail) tuples.
    """
    results = []
    if not os.path.exists(log_path):
        print(f"Error: Log file not found: {log_path}")
        sys.exit(1)

    with open(log_path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    pattern = re.compile(
        r'\[(PASSED|FAILED)\]\s+(.+?)(?:\s+\(Exit:\s*\d+\))?\n'
        r'(?:\s+Last line: (.+?)\n(?:\s+.+\n){0,5})?',
        re.MULTILINE
    )

    for match in pattern.finditer(content):
        status = match.group(1)
        script = match.group(2).strip()
        fail_type = ""
        fail_detail = ""
        if status == "FAILED":
            last_line = match.group(3) or ""
            subsequent_lines = []
            start = match.end()
            remaining = content[start:start + 500]
            lines = remaining.split('\n')
            for line in lines[:5]:
                stripped = line.strip()
                if stripped and not stripped.startswith('['):
                    subsequent_lines.append(stripped)
                else:
                    break
            all_last = [last_line] + subsequent_lines if last_line else subsequent_lines
            fail_type, fail_detail = classify_failure(all_last)
        results.append((script, status, fail_type, fail_detail))

    return results


def write_round_sheet(wb, sheet_name, results):
    """Write a round's detailed results to a new sheet."""
    ws = wb.create_sheet(sheet_name)

    headers = ["序号", "测试脚本", "结果", "失败类型", "失败详情"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, (script, status, fail_type, fail_detail) in enumerate(results, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=script).border = thin_border
        cell_status = ws.cell(row=row, column=3, value=status)
        cell_status.border = thin_border
        cell_status.alignment = Alignment(horizontal='center')
        if status == "PASSED":
            cell_status.fill = pass_fill
            cell_status.font = Font(color="006100")
        else:
            cell_status.fill = fail_fill
            cell_status.font = Font(color="9C0006")
        ws.cell(row=row, column=4, value=fail_type or None).border = thin_border
        ws.cell(row=row, column=5, value=fail_detail or None).border = thin_border

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 60

    passed = sum(1 for r in results if r[1] == "PASSED")
    failed = sum(1 for r in results if r[1] == "FAILED")

    summary_row = len(results) + 3
    ws.cell(row=summary_row, column=1, value="汇总").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)
    ws.cell(row=summary_row + 1, column=1, value="总数").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=len(results))
    ws.cell(row=summary_row + 2, column=1, value="通过").font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=2, value=passed).font = Font(color="006100", bold=True)
    ws.cell(row=summary_row + 3, column=1, value="失败").font = Font(bold=True)
    ws.cell(row=summary_row + 3, column=2, value=failed).font = Font(color="9C0006", bold=True)
    ws.cell(row=summary_row + 4, column=1, value="通过率").font = Font(bold=True)
    ws.cell(row=summary_row + 4, column=2, value=f"{passed / len(results) * 100:.0f}%").font = Font(bold=True, size=12)

    return passed, failed


def write_failure_summary_sheet(wb, results):
    """Write or update the failure classification summary sheet."""
    sheet_name = "失败分类汇总"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    fail_types = {}
    for script, status, fail_type, fail_detail in results:
        if status == "FAILED":
            key = fail_type if fail_type else "未知"
            fail_types.setdefault(key, []).append((script, fail_detail))

    headers = ["失败类型", "数量", "涉及算子"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for ft, items in fail_types.items():
        ws.cell(row=row, column=1, value=ft).border = thin_border
        ws.cell(row=row, column=2, value=len(items)).border = thin_border
        scripts = "; ".join([s for s, d in items])
        ws.cell(row=row, column=3, value=scripts).border = thin_border
        row += 1

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 80


def write_comparison_sheet(wb, round_sheets_info):
    """Write or update the cross-round comparison sheet.

    round_sheets_info: list of (sheet_name, {script: status}) dicts
    """
    sheet_name = "对比分析"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    all_scripts = sorted(set(
        s for _, results_dict in round_sheets_info for s in results_dict.keys()
    ))
    num_rounds = len(round_sheets_info)

    headers = ["序号", "测试脚本"]
    for sheet_name_r, _ in round_sheets_info:
        short_name = sheet_name_r.replace("Round ", "R")
        headers.append(short_name)
    if num_rounds >= 2:
        headers.append("变化(最新vs上一轮)")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 2
    for i, script in enumerate(all_scripts, 1):
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=script).border = thin_border

        for j, (_, results_dict) in enumerate(round_sheets_info):
            val = results_dict.get(script, "N/A")
            col_idx = 3 + j
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if val == "PASSED":
                cell.fill = pass_fill
                cell.font = Font(color="006100")
            elif val == "FAILED":
                cell.fill = fail_fill
                cell.font = Font(color="9C0006")

        if num_rounds >= 2:
            prev_status = round_sheets_info[-2][1].get(script, "N/A")
            curr_status = round_sheets_info[-1][1].get(script, "N/A")
            change = "无变化"
            if prev_status != curr_status:
                if prev_status == "FAILED" and curr_status == "PASSED":
                    change = "FIXED"
                elif prev_status == "PASSED" and curr_status == "FAILED":
                    change = "NEW FAIL"
                else:
                    change = "变化"

            col_change = 3 + num_rounds
            cell_change = ws.cell(row=row, column=col_change, value=change)
            cell_change.border = thin_border
            cell_change.alignment = Alignment(horizontal='center')
            if change == "FIXED":
                cell_change.fill = fixed_fill
                cell_change.font = Font(bold=True, color="006100")
            elif change == "NEW FAIL":
                cell_change.fill = new_fail_fill
                cell_change.font = Font(bold=True, color="FFFFFF")
            elif change == "变化":
                cell_change.fill = changed_fill
                cell_change.font = Font(bold=True)

        row += 1

    summary_row = row + 1
    ws.cell(row=summary_row, column=1, value="多轮对比汇总").font = Font(bold=True, size=14)
    ws.merge_cells(
        start_row=summary_row, start_column=1,
        end_row=summary_row, end_column=3 + num_rounds
    )

    summary_headers = ["指标"]
    for sheet_name_r, _ in round_sheets_info:
        short_name = sheet_name_r.replace("Round ", "R")
        summary_headers.append(short_name)
    if num_rounds >= 2:
        summary_headers.append("差值(最新-上一轮)")

    for col, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=summary_row + 1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = thin_border

    stats = []
    passed_counts = []
    failed_counts = []
    for _, results_dict in round_sheets_info:
        p = sum(1 for v in results_dict.values() if v == "PASSED")
        f = sum(1 for v in results_dict.values() if v == "FAILED")
        passed_counts.append(p)
        failed_counts.append(f)

    stats.append(("总数", [len(all_scripts)] * num_rounds,
                  [0] if num_rounds >= 2 else None))
    stats.append(("通过", passed_counts,
                  [passed_counts[-1] - passed_counts[-2]] if num_rounds >= 2 else None))
    stats.append(("失败", failed_counts,
                  [failed_counts[-1] - failed_counts[-2]] if num_rounds >= 2 else None))
    pass_rates = [f"{p/len(all_scripts)*100:.0f}%" for p in passed_counts]
    if num_rounds >= 2:
        rate_diff = f"{(passed_counts[-1]-passed_counts[-2])/len(all_scripts)*100:.1f}%"
        stats.append(("通过率", pass_rates, [rate_diff]))
    else:
        stats.append(("通过率", pass_rates, None))

    if num_rounds >= 2:
        fixed = sum(1 for s in all_scripts
                    if round_sheets_info[-2][1].get(s) == "FAILED"
                    and round_sheets_info[-1][1].get(s) == "PASSED")
        new_fail = sum(1 for s in all_scripts
                       if round_sheets_info[-2][1].get(s) == "PASSED"
                       and round_sheets_info[-1][1].get(s) == "FAILED")
        still_failing = failed_counts[-1] - new_fail
        stats.append(("修复(FIXED)", [str(fixed)], [str(fixed)]))
        stats.append(("新增失败(NEW FAIL)", [str(new_fail)], [str(new_fail)]))
        stats.append(("持续失败", [str(still_failing)], [str(-new_fail)]))

    for j, (label, values, diffs) in enumerate(stats):
        r = summary_row + 2 + j
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=1).border = thin_border
        for k, v in enumerate(values):
            ws.cell(row=r, column=2 + k, value=v).border = thin_border
        if diffs is not None:
            for k, d in enumerate(diffs):
                ws.cell(row=r, column=2 + num_rounds + k, value=d).border = thin_border

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 55
    for col_idx in range(3, 3 + num_rounds + (1 if num_rounds >= 2 else 0)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12


def get_next_round_number(wb):
    """Determine the next round number from existing sheet names."""
    max_round = 0
    for name in wb.sheetnames:
        match = re.match(r'Round (\d+)', name)
        if match:
            max_round = max(max_round, int(match.group(1)))
    return max_round + 1


def collect_round_data(wb):
    """Collect results data from all existing round sheets."""
    round_sheets_info = []
    for name in wb.sheetnames:
        match = re.match(r'Round (\d+)', name)
        if match:
            ws = wb[name]
            results_dict = {}
            for i in range(2, ws.max_row + 1):
                script = ws.cell(row=i, column=2).value
                status = ws.cell(row=i, column=3).value
                if script and status:
                    results_dict[script] = status
            round_sheets_info.append((name, results_dict))
    return round_sheets_info


def export_to_excel(log_path, backend, excel_path):
    """Main export function: parse log, create/update Excel."""
    results = parse_log(log_path)
    if not results:
        print("Warning: No test results parsed from log. Check log format.")
        return

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        round_num = get_next_round_number(wb)
        existing_round_data = collect_round_data(wb)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        round_num = 1
        existing_round_data = []

    sheet_name = f"Round {round_num} ({backend})"
    results_dict = {r[0]: r[1] for r in results}
    passed, failed = write_round_sheet(wb, sheet_name, results)

    all_round_data = existing_round_data + [(sheet_name, results_dict)]

    write_failure_summary_sheet(wb, results)

    if len(all_round_data) >= 2:
        write_comparison_sheet(wb, all_round_data)

    wb.save(excel_path)
    print(f"Excel saved to: {excel_path}")
    print(f"Sheet '{sheet_name}': {passed} passed, {failed} failed, "
          f"pass rate {passed / len(results) * 100:.0f}%")
    if len(all_round_data) >= 2:
        prev_results = all_round_data[-2][1]
        prev_passed = sum(1 for v in prev_results.values() if v == "PASSED")
        prev_failed = sum(1 for v in prev_results.values() if v == "FAILED")
        fixed = sum(1 for s in results_dict
                    if prev_results.get(s) == "FAILED" and results_dict[s] == "PASSED")
        new_fail = sum(1 for s in results_dict
                       if prev_results.get(s) == "PASSED" and results_dict[s] == "FAILED")
        print(f"Comparison vs {all_round_data[-2][0]}: "
              f"FIXED {fixed}, NEW FAIL {new_fail}, Still failing {failed - new_fail}")
    print(f"Total sheets: {wb.sheetnames}")


def main():
    parser = argparse.ArgumentParser(
        description="Export TileLang-Ascend test results to Excel"
    )
    parser.add_argument(
        "--log", required=True,
        help="Path to run_examples.sh output log file"
    )
    parser.add_argument(
        "--backend", required=True, choices=["auto", "ascendc", "pto"],
        help="Backend type used in the test run (auto, ascendc or pto)"
    )
    parser.add_argument(
        "--excel",
        default=os.path.join(os.getcwd(), "test_results.xlsx"),
        help="Path to Excel output file (default: test_results.xlsx in current dir)"
    )
    args = parser.parse_args()
    export_to_excel(args.log, args.backend, args.excel)


if __name__ == "__main__":
    main()